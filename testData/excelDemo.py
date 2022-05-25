import openpyxl

sheet = openpyxl.load_workbook("./testData.xlsx")

sheetname = sheet.active

Dict = {}

print(sheetname)  # <Worksheet "Sheet1">

print(sheetname.cell(row=2, column=1).value)  # varun

print(sheetname.max_row)  # 3
print(sheetname.max_column)  # 3


for r in range(2, sheetname.max_row + 1):
    if sheetname.cell(row=r, column=1).value == "varun":
        for c in range(1, sheetname.max_column + 1):
            Dict[sheetname.cell(row=1, column=c).value] = sheetname.cell(row=r, column=c).value

print(Dict)